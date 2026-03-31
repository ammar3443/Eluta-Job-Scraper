# scraper.py
import json
import os
import re
import sys
import time
import random
import argparse
from datetime import date
from difflib import SequenceMatcher
from urllib.robotparser import RobotFileParser

import yaml
import requests
from bs4 import BeautifulSoup
import anthropic
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

ELUTA_BASE = "https://www.eluta.ca"
ELUTA_SEARCH = f"{ELUTA_BASE}/search"

CATEGORY_COLORS = {
    "backend":     "B8CCE4",
    "frontend":    "C6EFCE",
    "fullstack":   "DDEBF7",
    "ai_ml":       "E2CFED",
    "firmware":    "FCE4D6",
    "cloud_devops":"DAEEF3",
    "mobile":      "FFF2CC",
    "data":        "EDEDED",
    "analyst":     "F2F2F2",
    "general_swe": "FFFFFF",
}

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )
}

# ---------------------------------------------------------------------------
# Config + Feedback
# ---------------------------------------------------------------------------

def load_config(path: str = "config.yaml") -> dict:
    try:
        with open(path, "r") as f:
            return yaml.safe_load(f)
    except FileNotFoundError:
        sys.exit(f"Error: config file not found: {path}")
    except yaml.YAMLError as exc:
        sys.exit(f"Error: invalid YAML in {path}: {exc}")


def load_feedback(path: str = "feedback.json") -> dict:
    if not os.path.exists(path):
        return {"decisions": [], "ambiguous_titles": []}
    with open(path, "r") as f:
        return json.load(f)


def save_feedback(feedback: dict, path: str = "feedback.json") -> None:
    with open(path, "w") as f:
        json.dump(feedback, f, indent=2)


def load_seen_ids(path: str = "seen_jobs.json") -> set[str]:
    """Load job IDs seen in previous runs."""
    if not os.path.exists(path):
        return set()
    with open(path, "r") as f:
        return set(json.load(f))


def save_seen_ids(seen_ids: set[str], path: str = "seen_jobs.json") -> None:
    with open(path, "w") as f:
        json.dump(list(seen_ids), f)


def _parse_days_ago(date_str: str) -> int:
    """Parse Eluta's relative date string to approximate number of days ago."""
    s = date_str.lower().strip()
    if not s or "today" in s or "hour" in s or "minute" in s or "just" in s:
        return 0
    if "yesterday" in s:
        return 1
    m = re.search(r"(\d+)\s*day", s)
    if m:
        return int(m.group(1))
    m = re.search(r"(\d+)\s*week", s)
    if m:
        return int(m.group(1)) * 7
    m = re.search(r"(\d+)\s*month", s)
    if m:
        return int(m.group(1)) * 30
    return 0  # unknown format — don't skip


# ---------------------------------------------------------------------------
# Hard Filter
# ---------------------------------------------------------------------------

def hard_filter(title: str, config: dict, ambiguous_titles: set) -> tuple[str, str | None]:
    """
    Returns (action, reason).
    action: "pass" | "filter" | "ambiguous"
    reason: description string if filtered, None if passed/ambiguous
    """
    title_lower = title.lower().strip()

    # Ambiguous list overrides all blocklists — goes straight to Claude
    if title_lower in ambiguous_titles:
        return ("ambiguous", None)

    filters = config["filters"]

    # Seniority blocklist — prepend a space so " lead" matches both
    # "Lead Engineer" (start of title) and "Team Lead" (mid-title)
    padded_title = " " + title_lower
    for term in filters["seniority_blocklist"]:
        if term.lower() in padded_title:
            return ("filter", f"seniority blocklist match: '{term.strip()}'")

    # Non-technical blocklist
    for term in filters["non_technical_blocklist"]:
        if term.lower() in title_lower:
            return ("filter", f"non-technical blocklist match: '{term.strip()}'")

    return ("pass", None)


# ---------------------------------------------------------------------------
# YOE Extractor
# ---------------------------------------------------------------------------

def _categorize_yoe(years: int) -> str:
    if years <= 1:
        return "0-1"
    elif years <= 3:
        return "2-3"
    elif years < 5:
        return "4-5"
    else:
        return "5+"


def extract_yoe(text: str) -> str:
    """Extract years-of-experience requirement from job description text."""
    text_lower = text.lower()

    # New grad / intern / entry level first
    entry_patterns = ["new grad", "new graduate", "entry level", "entry-level",
                      "internship", "co-op", "coop", "no experience"]
    if any(p in text_lower for p in entry_patterns):
        return "0-1"

    # "3-5 years" or "3 to 5 years"
    range_match = re.search(
        r"(\d+)\s*(?:-|to)\s*(\d+)\s*\+?\s*years?", text_lower
    )
    if range_match:
        return _categorize_yoe(int(range_match.group(1)))

    # "5+ years experience" or "5 years of experience"
    single_match = re.search(
        r"(\d+)\s*\+?\s*years?\s*(?:of\s+)?(?:experience|exp)", text_lower
    )
    if single_match:
        return _categorize_yoe(int(single_match.group(1)))

    # "experience: 3 years" style
    exp_colon = re.search(r"experience[:\s]+(\d+)\s*\+?\s*years?", text_lower)
    if exp_colon:
        return _categorize_yoe(int(exp_colon.group(1)))

    return "unknown"


# ---------------------------------------------------------------------------
# Keyword Classifier
# ---------------------------------------------------------------------------

def keyword_classify(title: str, categories: dict) -> str | None:
    """
    Match job title against category keyword lists.
    Returns category name or None if no match (→ Claude).
    general_swe has no keywords intentionally — never keyword-matched.
    """
    title_lower = title.lower()
    for category, keywords in categories.items():
        if not keywords:
            continue
        for kw in keywords:
            if kw.lower() in title_lower:
                return category
    return None


# ---------------------------------------------------------------------------
# Feedback Lookup
# ---------------------------------------------------------------------------

_FUZZY_THRESHOLD = 0.85


def feedback_lookup(title: str, feedback: dict) -> dict | None:
    """
    Check feedback decisions for a matching title.
    Returns the decision dict {title, relevant, category, reason} or None.
    Tries exact match first, then fuzzy match with ratio >= 0.85.
    """
    title_lower = title.lower().strip()
    decisions = feedback.get("decisions", [])

    # Exact match
    for d in decisions:
        if d["title"].lower().strip() == title_lower:
            return d

    # Fuzzy match
    best_ratio = 0.0
    best_decision = None
    for d in decisions:
        ratio = SequenceMatcher(None, title_lower, d["title"].lower().strip()).ratio()
        if ratio > best_ratio:
            best_ratio = ratio
            best_decision = d

    if best_ratio >= _FUZZY_THRESHOLD:
        return best_decision

    return None


# ---------------------------------------------------------------------------
# Claude Prompt + Response Parser
# ---------------------------------------------------------------------------

_CATEGORIES_LIST = (
    "backend, frontend, fullstack, ai_ml, firmware, "
    "cloud_devops, mobile, data, analyst, general_swe"
)


def build_claude_prompt(title: str, jd_text: str, feedback: dict, config: dict) -> str:
    max_examples = config["classifier"]["max_few_shot_examples"]
    decisions = feedback.get("decisions", [])[-max_examples:]

    examples_block = ""
    if decisions:
        lines = []
        for d in decisions:
            if d["relevant"]:
                line = f'- "{d["title"]}" → relevant, category: {d["category"]}'
                if d.get("reason"):
                    line += f' [{d["reason"]}]'
            else:
                line = f'- "{d["title"]}" → NOT relevant'
                if d.get("reason"):
                    line += f' [reason: {d["reason"]}]'
            lines.append(line)
        examples_block = "Past decisions (use these as guidance):\n" + "\n".join(lines) + "\n\n"

    return (
        "You are classifying job postings for a software engineering job search.\n\n"
        f"{examples_block}"
        f"Categories: {_CATEGORIES_LIST}\n\n"
        "Rules:\n"
        "- When in doubt about category, use general_swe — do not miss a relevant posting\n"
        "- Non-technical roles (civil, mechanical, industrial, trades) → relevant: false\n"
        "- Read the full job description for technology keywords (AWS, Python, Terraform, etc.)\n"
        "- Return ONLY JSON, no explanation\n\n"
        f"Job title: {title}\n"
        f"Job description: {jd_text}\n\n"
        'Reply with exactly:\n'
        '{"relevant": true/false, "category": "<category>", "confidence": 0.0-1.0, "yoe": "0-1|2-3|4-5|5+|unknown"}'
    )


def parse_claude_response(raw: str) -> dict:
    """
    Extract JSON from Claude's response text.
    Returns a safe fallback dict with low confidence if parsing fails.
    """
    # Try to find a JSON object in the response
    match = re.search(r"\{[^{}]+\}", raw, re.DOTALL)
    if match:
        try:
            data = json.loads(match.group())
            return {
                "relevant": bool(data.get("relevant", False)),
                "category": data.get("category", "general_swe"),
                "confidence": float(data.get("confidence", 0.0)),
                "yoe": data.get("yoe", "unknown"),
            }
        except (json.JSONDecodeError, ValueError):
            print(f"Warning: Claude response matched JSON pattern but failed to parse: {match.group()[:100]}")
    else:
        print(f"Warning: Could not find JSON in Claude response: {raw[:100]}")

    # Fallback: unparseable response → flag for review
    return {"relevant": False, "category": None, "confidence": 0.0, "yoe": "unknown"}


# ---------------------------------------------------------------------------
# Claude Classifier
# ---------------------------------------------------------------------------

def claude_classify(title: str, jd_text: str, feedback: dict, config: dict) -> dict:
    """
    Call Claude Haiku to classify a borderline job.
    Returns dict with keys: relevant, category, confidence, yoe, flagged_for_review.
    Requires ANTHROPIC_API_KEY environment variable.
    """
    client = anthropic.Anthropic()  # reads ANTHROPIC_API_KEY from env
    prompt = build_claude_prompt(title, jd_text, feedback, config)

    response = client.messages.create(
        model=config["classifier"]["claude_model"],
        max_tokens=100,
        messages=[{"role": "user", "content": prompt}],
    )

    raw = response.content[0].text
    result = parse_claude_response(raw)
    threshold = config["classifier"]["confidence_threshold"]
    result["flagged_for_review"] = result["confidence"] < threshold
    return result


# ---------------------------------------------------------------------------
# Scraper
# ---------------------------------------------------------------------------

def _check_robots(config: dict) -> bool:
    """Returns True if scraping is allowed. Exits if robots.txt disallows and respect=True."""
    if not config["scraper"].get("respect_robots_txt", True):
        return True
    rp = RobotFileParser()
    rp.set_url(f"{ELUTA_BASE}/robots.txt")
    rp.read()
    allowed = rp.can_fetch(HEADERS["User-Agent"], ELUTA_SEARCH)
    if not allowed:
        print("robots.txt disallows scraping eluta.ca. Exiting.")
        sys.exit(1)
    return True


def _polite_delay(config: dict) -> None:
    delay = random.uniform(
        config["scraper"]["delay_min"],
        config["scraper"]["delay_max"],
    )
    time.sleep(delay)


def _extract_job_id(slug: str) -> str:
    """Extract the hash ID from a slug like 'spl/backend-dev-<32hexchars>?imo=1'."""
    match = re.search(r"([a-f0-9]{32})", slug)
    return match.group(1) if match else slug.split("?")[0]


def fetch_results_page(page: int, query: str, config: dict) -> list[dict]:
    """
    Fetch one search results page from Eluta.
    Returns list of job dicts: {title, company, snippet, date_posted, job_id, slug, url}
    """
    # Eluta uses "pg" for pagination; page 1 has no pg parameter, pages 2+ use pg=N
    params = {"q": query}
    if page > 1:
        params["pg"] = page
    _polite_delay(config)
    resp = requests.get(ELUTA_SEARCH, params=params, headers=HEADERS, timeout=30)
    resp.raise_for_status()  # surface 403/429/5xx immediately instead of silently parsing error HTML
    soup = BeautifulSoup(resp.text, "html.parser")

    jobs = []
    for card in soup.find_all(class_="organic-job"):
        title_tag = card.find("a", class_="lk-job-title")
        company_tag = card.find("a", class_="lk-employer")
        desc_tag = card.find("span", class_="description")
        date_tag = card.find("a", class_="lastseen")
        slug = card.get("data-url", "")

        if not title_tag or not slug:
            continue

        job_id = _extract_job_id(slug)
        jobs.append({
            "title": title_tag.get_text(strip=True),
            "company": company_tag.get_text(strip=True) if company_tag else "",
            "snippet": desc_tag.get_text(strip=True) if desc_tag else "",
            "date_posted": date_tag.get_text(strip=True) if date_tag else "",
            "job_id": job_id,
            "slug": slug,
            "url": f"{ELUTA_BASE}/{slug.split('?')[0]}",  # clean URL without ?imo= param
        })

    if not jobs:
        print(f"  [debug] Page {page} returned no jobs. URL: {resp.url}")
        print(f"  [debug] Response snippet: {resp.text[:500]}")

    return jobs


def fetch_full_jd(slug: str, config: dict) -> str:
    """
    Fetch the full job description from the /spl/ page.
    Returns plain text of the description, or empty string on failure.
    """
    # slug may be "spl/job-title-hash?imo=N" — build full URL
    url = f"{ELUTA_BASE}/{slug}" if not slug.startswith("http") else slug
    _polite_delay(config)
    resp = requests.get(url, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    desc = soup.find(class_="short-text")
    if not desc:
        # Fallback: try description div
        desc = soup.find("div", class_="description")
    if not desc:
        return ""
    return desc.get_text(separator=" ", strip=True)


# ---------------------------------------------------------------------------
# Main Pipeline
# ---------------------------------------------------------------------------

def classify_job(job: dict, jd_text: str, feedback: dict, config: dict, is_ambiguous: bool = False) -> dict:
    """
    Run the Flow D classifier on a single job.
    is_ambiguous: True when the title came from the ambiguous list — skip straight to Claude.
    Returns job dict enriched with: category, confidence, flagged_for_review, yoe_required.
    """
    title = job["title"]

    if not is_ambiguous:
        # Step 1: feedback lookup
        fb_decision = feedback_lookup(title, feedback)
        if fb_decision:
            if not fb_decision["relevant"]:
                # Previously confirmed as not relevant — filter it
                return {**job, "relevant": False, "category": None,
                        "confidence": None, "flagged_for_review": False, "yoe_required": "unknown"}
            return {
                **job,
                "relevant": True,
                "category": fb_decision["category"] or "general_swe",
                "confidence": None,  # blank for feedback-matched jobs (spec)
                "flagged_for_review": False,
                "yoe_required": extract_yoe(jd_text),
            }

        # Step 2: keyword match
        kw_category = keyword_classify(title, config["categories"])
        if kw_category:
            return {
                **job,
                "relevant": True,
                "category": kw_category,
                "confidence": None,  # blank for keyword-matched jobs (spec)
                "flagged_for_review": False,
                "yoe_required": extract_yoe(jd_text),
            }

    # Step 3: Claude
    claude_result = claude_classify(title, jd_text, feedback, config)
    yoe = extract_yoe(jd_text)  # compute once; fall back to Claude's yoe if regex found nothing
    return {
        **job,
        "category": claude_result["category"] or "general_swe",
        "confidence": claude_result["confidence"],
        "flagged_for_review": claude_result["flagged_for_review"],
        "relevant": claude_result["relevant"],
        "yoe_required": yoe if yoe != "unknown" else claude_result.get("yoe", "unknown"),
    }


def run_scrape(config: dict, feedback: dict, seen_ids: set[str] | None = None) -> tuple[list, list, list, int, int, str | None]:
    """
    Main scrape loop. Returns (accepted, review, filtered, duplicate_count, pages_scraped).
    - accepted: classified jobs above confidence threshold
    - review: borderline jobs flagged for human review
    - filtered: hard-filtered jobs (non-technical / seniority)
    - duplicate_count: jobs skipped due to duplicate job_id in this run
    - pages_scraped: number of result pages fetched

    seen_ids: set of job IDs already processed in previous runs (passed in from main).
    """
    _check_robots(config)

    eluta_cfg = config["sites"]["eluta"]
    query = eluta_cfg["query"]
    max_pages = eluta_cfg["max_pages"]
    cutoff_days = config["scraper"].get("cutoff_days", 3)

    if seen_ids is None:
        seen_ids = set()
    accepted: list[dict] = []
    review: list[dict] = []
    filtered: list[dict] = []
    duplicate_count = 0
    pages_scraped = 0
    ambiguous_titles = {t.lower() for t in feedback.get("ambiguous_titles", [])}

    network_error: str | None = None

    for page in range(1, max_pages + 1):
        try:
            page_jobs = fetch_results_page(page, query, config)
        except requests.RequestException as exc:
            network_error = str(exc)
            print(f"\n  Network error on page {page}: {exc}")
            print("  Saving results collected so far.")
            break

        if not page_jobs:
            break  # No more results

        # Date cutoff: if every job on this page is older than cutoff_days, stop.
        # Filter to only jobs within the cutoff before processing.
        if cutoff_days:
            page_jobs = [j for j in page_jobs
                         if _parse_days_ago(j.get("date_posted", "")) <= cutoff_days]
            if not page_jobs:
                print(f"\n  Reached {cutoff_days}-day cutoff at page {page}, stopping.")
                break

        pages_scraped += 1
        print(f"\r  Page {pages_scraped}...", end="", flush=True)

        for job in page_jobs:
            jid = job["job_id"]
            if jid in seen_ids:
                duplicate_count += 1
                continue
            seen_ids.add(jid)

            action, reason = hard_filter(job["title"], config, ambiguous_titles)

            if action == "filter":
                filtered.append({**job, "filter_reason": reason})
                continue

            # Fetch full JD for all non-filtered jobs
            try:
                jd_text = fetch_full_jd(job["slug"], config)
            except requests.RequestException as exc:
                print(f"\n  Network error fetching JD for '{job['title']}': {exc} — skipping job.")
                continue

            classified = classify_job(job, jd_text, feedback, config, is_ambiguous=(action == "ambiguous"))

            if not classified["relevant"]:
                filtered.append({**job, "filter_reason": "Claude: not relevant"})
                continue

            if classified.get("flagged_for_review"):
                review.append(classified)
            else:
                accepted.append(classified)

    print()  # newline after the inline page counter
    return accepted, review, filtered, duplicate_count, pages_scraped, network_error


# ---------------------------------------------------------------------------
# XLSX Export
# ---------------------------------------------------------------------------

_ACCEPTED_COLUMNS = ["job_id", "title", "company", "date_posted", "category",
                     "yoe_required", "url", "confidence"]

_REVIEW_COLUMNS = _ACCEPTED_COLUMNS + ["confirm", "reason"]

_HEADER_FILL = PatternFill("solid", fgColor="2F4F4F")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _apply_xlsx_formatting(ws, columns: list[str]) -> None:
    """Apply header formatting, auto-filter, and freeze top row."""
    # Header row
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(1, col_idx)
        cell.value = col_name
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Auto-filter on all columns
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    # Freeze top row
    ws.freeze_panes = "A2"

    # Column widths
    widths = {"title": 40, "company": 25, "url": 50, "category": 15,
              "yoe_required": 12, "date_posted": 15, "confidence": 12,
              "job_id": 34, "confirm": 10, "reason": 40}
    for col_idx, col_name in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 15)


def _write_job_row(ws, row_idx: int, job: dict, columns: list[str]) -> None:
    """Write a single job row with category color-coding and URL hyperlink."""
    fill_color = CATEGORY_COLORS.get(job.get("category", "general_swe"), "FFFFFF")
    row_fill = PatternFill("solid", fgColor=fill_color)

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row_idx, col_idx)
        value = job.get(col_name, "")

        if col_name == "url" and value:
            cell.hyperlink = value
            cell.value = value
            cell.font = Font(color="0563C1", underline="single")
        elif col_name == "confidence":
            cell.value = f"{value:.0%}" if isinstance(value, (int, float)) else ""
        else:
            cell.value = value if value is not None else ""

        cell.fill = row_fill


def write_accepted_xlsx(jobs: list[dict], filepath: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"
    _apply_xlsx_formatting(ws, _ACCEPTED_COLUMNS)
    for i, job in enumerate(jobs, start=2):
        _write_job_row(ws, i, job, _ACCEPTED_COLUMNS)
    wb.save(filepath)


def write_review_xlsx(jobs: list[dict], filepath: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Review"
    _apply_xlsx_formatting(ws, _REVIEW_COLUMNS)
    for i, job in enumerate(jobs, start=2):
        _write_job_row(ws, i, job, _REVIEW_COLUMNS)
        # Leave confirm and reason blank for user to fill
    wb.save(filepath)


# ---------------------------------------------------------------------------
# Filtered JSON Export
# ---------------------------------------------------------------------------

_FILTERED_FIELDS = ["job_id", "title", "company", "date_posted", "snippet",
                    "filter_reason", "url"]


def write_filtered_json(jobs: list[dict], filepath: str) -> None:
    records = [{k: job.get(k, "") for k in _FILTERED_FIELDS} for job in jobs]
    with open(filepath, "w") as f:
        json.dump(records, f, indent=2)


# ---------------------------------------------------------------------------
# Feedback Ingester
# ---------------------------------------------------------------------------

def _ingest_from_review_xlsx(filepath: str, feedback: dict) -> dict:
    """Process a review XLSX file. Reads confirm/reason columns filled in by user."""
    wb = load_workbook(filepath)
    ws = wb.active

    # Find column indices from header row
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    required = {"title", "category", "confirm", "reason"}
    if not required.issubset(headers.keys()):
        print(f"Warning: review file missing columns {required - set(headers.keys())}")
        return feedback

    for row in range(2, ws.max_row + 1):
        confirm_val = ws.cell(row, headers["confirm"]).value
        if not confirm_val:
            continue  # User left this row blank — skip
        title = ws.cell(row, headers["title"]).value or ""
        category = ws.cell(row, headers["category"]).value or "general_swe"
        reason = ws.cell(row, headers["reason"]).value or ""
        confirmed = str(confirm_val).strip().lower() in ("yes", "y", "true", "1")

        feedback["decisions"].append({
            "title": title,
            "relevant": confirmed,
            "category": category if confirmed else None,
            "reason": reason,
            "source": "review",
        })

    return feedback


def _ingest_from_dispute_json(filepath: str, feedback: dict) -> dict:
    """Process a filtered JSON file. Only processes entries with dispute: true."""
    with open(filepath) as f:
        jobs = json.load(f)

    existing_lower = {t.lower() for t in feedback["ambiguous_titles"]}
    for job in jobs:
        if not job.get("dispute"):
            continue
        title = job.get("title", "")
        reason = job.get("reason", "")

        # Add to ambiguous list (bypasses hard filter next run → goes to Claude)
        title_lower = title.lower().strip()
        if title_lower not in existing_lower:
            feedback["ambiguous_titles"].append(title_lower)
            existing_lower.add(title_lower)

        # Add as positive few-shot example for Claude
        feedback["decisions"].append({
            "title": title,
            "relevant": True,
            "category": "general_swe",  # Claude will re-classify properly next run
            "reason": reason,
            "source": "dispute",
        })

    return feedback


def ingest_feedback(filepath: str, feedback: dict) -> dict:
    """
    Auto-detect file type by extension and ingest feedback.
    .xlsx → review file (confirm/reject borderline jobs)
    .json → dispute file (contest hard-filtered jobs)
    """
    if filepath.endswith(".xlsx"):
        return _ingest_from_review_xlsx(filepath, feedback)
    elif filepath.endswith(".json"):
        return _ingest_from_dispute_json(filepath, feedback)
    else:
        print(f"Unknown file type: {filepath}. Expected .xlsx or .json")
        return feedback


# ---------------------------------------------------------------------------
# Terminal Summary
# ---------------------------------------------------------------------------

def print_summary(accepted: list, review: list, filtered: list,
                  duplicate_count: int, pages_scraped: int,
                  out_path: str, elapsed: float) -> None:
    total = len(accepted) + len(review) + len(filtered) + duplicate_count
    minutes, seconds = divmod(int(elapsed), 60)
    elapsed_str = f"{minutes}m {seconds}s" if minutes else f"{seconds}s"
    print()
    print(f"Scraped {total} jobs across {pages_scraped} pages in {elapsed_str}")
    print(f"  Accepted:        {len(accepted)}")
    print(f"  Hard filtered:   {len(filtered)}  → check output/filtered_{date.today()}.json")
    print(f"  Flagged review:  {len(review)}  → check output/review_{date.today()}.xlsx")
    print(f"  Duplicate skip:  {duplicate_count}")
    print()
    print(f"Saved → {out_path}")
    print()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Eluta.ca Job Scraper")
    parser.add_argument(
        "--ingest-feedback", metavar="FILE",
        help="Ingest a completed review.xlsx or disputed filtered.json into feedback.json"
    )
    parser.add_argument("--config", default="config.yaml", help="Path to config file")
    parser.add_argument("--feedback", default="feedback.json", help="Path to feedback file")
    args = parser.parse_args()

    config = load_config(args.config)
    feedback = load_feedback(args.feedback)

    # Feedback ingestion mode
    if args.ingest_feedback:
        print(f"Ingesting feedback from {args.ingest_feedback}...")
        before_count = len(feedback.get("decisions", []))  # capture before in-place mutation
        updated = ingest_feedback(args.ingest_feedback, feedback)
        save_feedback(updated, args.feedback)
        new_decisions = len(updated["decisions"]) - before_count
        print(f"Done. {new_decisions} new decision(s) added to {args.feedback}.")
        return

    # Scrape mode
    os.makedirs("output", exist_ok=True)
    today = date.today()

    print(f"Starting Eluta scrape: '{config['sites']['eluta']['query']}'")
    print(f"Max pages: {config['sites']['eluta']['max_pages']}")
    print()
    start_time = time.time()

    seen_ids = load_seen_ids()
    try:
        accepted, review, filtered, dup_count, pages_scraped, network_error = run_scrape(config, feedback, seen_ids)
    except anthropic.AuthenticationError:
        sys.exit("Error: ANTHROPIC_API_KEY is missing or invalid. Set it with: export ANTHROPIC_API_KEY=...")
    save_seen_ids(seen_ids)

    accepted_path = f"output/eluta_{today}.xlsx"
    review_path = f"output/review_{today}.xlsx"
    filtered_path = f"output/filtered_{today}.json"

    write_accepted_xlsx(accepted, accepted_path)
    if review:
        write_review_xlsx(review, review_path)
    if filtered:
        write_filtered_json(filtered, filtered_path)

    if network_error:
        print(f"\nWarning: run ended early due to network error: {network_error}")
        print("Partial results saved.")

    print_summary(accepted, review, filtered, dup_count, pages_scraped, accepted_path, time.time() - start_time)


if __name__ == "__main__":
    main()
