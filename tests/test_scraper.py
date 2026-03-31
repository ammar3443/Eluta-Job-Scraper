# tests/test_scraper.py
import json
import os
import pytest
import yaml
from openpyxl import load_workbook
from unittest.mock import MagicMock, mock_open, patch


def test_load_config_returns_dict(tmp_path):
    cfg = {
        "sites": {"eluta": {"enabled": True, "query": "software engineer", "max_pages": 5}},
        "categories": {"backend": ["backend developer"], "general_swe": []},
        "filters": {"seniority_blocklist": [" lead"], "non_technical_blocklist": ["millwright"]},
        "classifier": {"confidence_threshold": 0.60, "claude_model": "claude-haiku-4-5-20251001", "max_few_shot_examples": 15},
        "scraper": {"delay_min": 1, "delay_max": 2, "respect_robots_txt": False},
    }
    config_file = tmp_path / "config.yaml"
    config_file.write_text(yaml.dump(cfg))
    from scraper import load_config
    result = load_config(str(config_file))
    assert result["classifier"]["confidence_threshold"] == 0.60
    assert result["sites"]["eluta"]["query"] == "software engineer"


def test_load_feedback_returns_structure(tmp_path):
    fb = {"decisions": [], "ambiguous_titles": []}
    fb_file = tmp_path / "feedback.json"
    fb_file.write_text(json.dumps(fb))
    from scraper import load_feedback
    result = load_feedback(str(fb_file))
    assert "decisions" in result
    assert "ambiguous_titles" in result


def test_load_feedback_missing_file_returns_empty():
    from scraper import load_feedback
    result = load_feedback("/nonexistent/feedback.json")
    assert result == {"decisions": [], "ambiguous_titles": []}


def test_save_feedback_writes_json(tmp_path):
    fb = {"decisions": [{"title": "Backend Dev", "relevant": True}], "ambiguous_titles": []}
    fb_file = tmp_path / "feedback.json"
    from scraper import save_feedback
    save_feedback(fb, str(fb_file))
    written = json.loads(fb_file.read_text())
    assert written["decisions"][0]["title"] == "Backend Dev"


def test_hard_filter_blocks_seniority():
    from scraper import hard_filter
    config = {
        "filters": {
            "seniority_blocklist": [" lead", "team lead", "engineering manager"],
            "non_technical_blocklist": ["millwright"],
        }
    }
    action, reason = hard_filter("Engineering Lead", config, [])
    assert action == "filter"
    assert "seniority" in reason.lower()


def test_hard_filter_blocks_non_technical():
    from scraper import hard_filter
    config = {
        "filters": {
            "seniority_blocklist": [" lead"],
            "non_technical_blocklist": ["millwright", "electrician"],
        }
    }
    action, reason = hard_filter("Millwright Technician", config, [])
    assert action == "filter"
    assert "non-technical" in reason.lower()


def test_hard_filter_passes_clean_title():
    from scraper import hard_filter
    config = {
        "filters": {
            "seniority_blocklist": [" lead"],
            "non_technical_blocklist": ["millwright"],
        }
    }
    action, reason = hard_filter("Backend Developer", config, [])
    assert action == "pass"
    assert reason is None


def test_hard_filter_ambiguous_list_overrides_blocklist():
    from scraper import hard_filter
    config = {
        "filters": {
            "seniority_blocklist": [" lead"],
            "non_technical_blocklist": ["civil engineer"],
        }
    }
    # "controls engineer" was disputed, now in ambiguous list
    action, reason = hard_filter("Controls Engineer", config, ["controls engineer"])
    assert action == "ambiguous"


def test_hard_filter_case_insensitive():
    from scraper import hard_filter
    config = {
        "filters": {
            "seniority_blocklist": [" lead"],
            "non_technical_blocklist": ["millwright"],
        }
    }
    action, _ = hard_filter("MILLWRIGHT OPERATOR", config, [])
    assert action == "filter"


# ---------------------------------------------------------------------------
# TASK 4: YOE Extractor Tests
# ---------------------------------------------------------------------------

def test_extract_yoe_range():
    from scraper import extract_yoe
    # Lower bound (3) passed to _categorize_yoe → "2-3"
    assert extract_yoe("We require 3-5 years of experience in Python.") == "2-3"


def test_extract_yoe_single():
    from scraper import extract_yoe
    assert extract_yoe("Minimum 2 years of experience required.") == "2-3"


def test_extract_yoe_plus():
    from scraper import extract_yoe
    assert extract_yoe("7+ years experience with distributed systems.") == "5+"


def test_extract_yoe_new_grad():
    from scraper import extract_yoe
    assert extract_yoe("This is a new grad position, no experience needed.") == "0-1"


def test_extract_yoe_internship():
    from scraper import extract_yoe
    assert extract_yoe("Summer internship for computer science students.") == "0-1"


def test_extract_yoe_entry_level():
    from scraper import extract_yoe
    assert extract_yoe("Entry-level software engineer role.") == "0-1"


def test_extract_yoe_unknown():
    from scraper import extract_yoe
    assert extract_yoe("Join our team and work on exciting projects.") == "unknown"


def test_extract_yoe_one_year():
    from scraper import extract_yoe
    assert extract_yoe("1 year of experience with React.") == "0-1"


# ---------------------------------------------------------------------------
# TASK 5: Keyword Classifier Tests
# ---------------------------------------------------------------------------

def _make_categories():
    return {
        "backend": ["backend developer", "api engineer"],
        "frontend": ["frontend developer", "ui engineer"],
        "ai_ml": ["ml engineer", "machine learning engineer"],
        "firmware": ["firmware engineer", "embedded software engineer"],
        "cloud_devops": ["cloud engineer", "devops engineer"],
        "general_swe": [],
    }


def test_keyword_classify_exact_match():
    from scraper import keyword_classify
    assert keyword_classify("Backend Developer", _make_categories()) == "backend"


def test_keyword_classify_case_insensitive():
    from scraper import keyword_classify
    assert keyword_classify("FRONTEND DEVELOPER", _make_categories()) == "frontend"


def test_keyword_classify_partial_match():
    from scraper import keyword_classify
    # "Senior ML Engineer" contains "ml engineer"
    assert keyword_classify("Senior ML Engineer", _make_categories()) == "ai_ml"


def test_keyword_classify_no_match_returns_none():
    from scraper import keyword_classify
    assert keyword_classify("Technical Specialist", _make_categories()) is None


def test_keyword_classify_general_swe_never_matches():
    # general_swe has empty keywords — always falls through to Claude
    from scraper import keyword_classify
    assert keyword_classify("Software Engineer", _make_categories()) is None


# ---------------------------------------------------------------------------
# TASK 6: Feedback Lookup Tests
# ---------------------------------------------------------------------------

def _make_feedback():
    return {
        "decisions": [
            {"title": "Platform Engineer", "relevant": True, "category": "cloud_devops", "reason": "confirmed"},
            {"title": "Process Control Engineer", "relevant": False, "category": None, "reason": "industrial"},
        ],
        "ambiguous_titles": ["controls engineer"]
    }


def test_feedback_lookup_exact_match():
    from scraper import feedback_lookup
    result = feedback_lookup("Platform Engineer", _make_feedback())
    assert result["relevant"] is True
    assert result["category"] == "cloud_devops"


def test_feedback_lookup_case_insensitive():
    from scraper import feedback_lookup
    result = feedback_lookup("platform engineer", _make_feedback())
    assert result is not None
    assert result["relevant"] is True


def test_feedback_lookup_fuzzy_match():
    from scraper import feedback_lookup
    # "Platform Engineer II" is close enough to "Platform Engineer"
    result = feedback_lookup("Platform Engineer II", _make_feedback())
    assert result is not None
    assert result["category"] == "cloud_devops"


def test_feedback_lookup_no_match_returns_none():
    from scraper import feedback_lookup
    result = feedback_lookup("iOS Developer", _make_feedback())
    assert result is None


def test_feedback_lookup_fuzzy_below_threshold_returns_none():
    from scraper import feedback_lookup
    # "Mechanical Engineer" is not similar enough to any stored title
    result = feedback_lookup("Mechanical Engineer", _make_feedback())
    assert result is None


# ---------------------------------------------------------------------------
# TASK 7: Claude Prompt Builder and Response Parser Tests
# ---------------------------------------------------------------------------

def test_build_claude_prompt_includes_title_and_jd():
    from scraper import build_claude_prompt
    feedback = {"decisions": [], "ambiguous_titles": []}
    config = {"classifier": {"max_few_shot_examples": 15}}
    prompt = build_claude_prompt("Platform Engineer", "We use AWS and Terraform.", feedback, config)
    assert "Platform Engineer" in prompt
    assert "AWS" in prompt
    assert "Terraform" in prompt


def test_build_claude_prompt_includes_few_shot_examples():
    from scraper import build_claude_prompt
    feedback = {
        "decisions": [
            {"title": "Embedded C Developer", "relevant": True, "category": "firmware", "reason": "PLC work"},
            {"title": "Civil Engineer", "relevant": False, "category": None, "reason": "not software"},
        ],
        "ambiguous_titles": []
    }
    config = {"classifier": {"max_few_shot_examples": 15}}
    prompt = build_claude_prompt("Some Job", "Some description.", feedback, config)
    assert "Embedded C Developer" in prompt
    assert "Civil Engineer" in prompt


def test_build_claude_prompt_caps_few_shot_examples():
    from scraper import build_claude_prompt
    decisions = [{"title": f"Job {i}", "relevant": True, "category": "backend", "reason": ""} for i in range(20)]
    feedback = {"decisions": decisions, "ambiguous_titles": []}
    config = {"classifier": {"max_few_shot_examples": 5}}
    prompt = build_claude_prompt("New Job", "Description.", feedback, config)
    # Only last 5 decisions should appear
    assert "Job 19" in prompt
    assert "Job 0" not in prompt


def test_parse_claude_response_valid_json():
    from scraper import parse_claude_response
    raw = '{"relevant": true, "category": "backend", "confidence": 0.92, "yoe": "2-3"}'
    result = parse_claude_response(raw)
    assert result["relevant"] is True
    assert result["category"] == "backend"
    assert result["confidence"] == 0.92
    assert result["yoe"] == "2-3"


def test_parse_claude_response_json_in_text():
    from scraper import parse_claude_response
    raw = 'Here is my answer: {"relevant": false, "category": null, "confidence": 0.95, "yoe": "unknown"} done.'
    result = parse_claude_response(raw)
    assert result["relevant"] is False


def test_parse_claude_response_malformed_returns_low_confidence():
    from scraper import parse_claude_response
    result = parse_claude_response("Sorry I cannot classify this.")
    assert result["confidence"] < 0.60
    assert result["relevant"] is False


# ---------------------------------------------------------------------------
# TASK 8: Claude Classifier Tests
# ---------------------------------------------------------------------------

def test_claude_classify_returns_classification():
    from scraper import claude_classify
    feedback = {"decisions": [], "ambiguous_titles": []}
    config = {
        "classifier": {
            "confidence_threshold": 0.60,
            "claude_model": "claude-haiku-4-5-20251001",
            "max_few_shot_examples": 15,
        }
    }
    mock_response = MagicMock()
    mock_response.content = [MagicMock(text='{"relevant": true, "category": "backend", "confidence": 0.93, "yoe": "2-3"}')]

    with patch("scraper.anthropic.Anthropic") as MockClient:
        MockClient.return_value.messages.create.return_value = mock_response
        result = claude_classify("Backend Developer", "Python, Django, REST APIs.", feedback, config)

    assert result["relevant"] is True
    assert result["category"] == "backend"
    assert result["confidence"] == 0.93
    assert result["yoe"] == "2-3"
    assert result["flagged_for_review"] is False


def test_claude_classify_flags_low_confidence():
    from scraper import claude_classify
    feedback = {"decisions": [], "ambiguous_titles": []}
    config = {
        "classifier": {
            "confidence_threshold": 0.60,
            "claude_model": "claude-haiku-4-5-20251001",
            "max_few_shot_examples": 15,
        }
    }
    mock_response = MagicMock()
    mock_response.content = [MagicMock(text='{"relevant": true, "category": "general_swe", "confidence": 0.45, "yoe": "unknown"}')]

    with patch("scraper.anthropic.Anthropic") as MockClient:
        MockClient.return_value.messages.create.return_value = mock_response
        result = claude_classify("Technical Specialist", "Some job.", feedback, config)

    assert result["flagged_for_review"] is True
    assert result["confidence"] == 0.45


# ---------------------------------------------------------------------------
# TASK 9: Scraper Tests
# ---------------------------------------------------------------------------

SAMPLE_RESULTS_HTML = """
<html><body>
<div class="organic-job odd" data-url="spl/backend-developer-d46b145bbcc78f3ebfdc6d584e74f6e7?imo=1">
  <h2 class="title">
    <a class="lk-job-title" href="#!">Backend Developer</a>
  </h2>
  <a class="employer lk-employer" href="#!">Acme Corp</a>
  <span class="description">We build APIs with Python and Django.</span>
  <a class="lk lastseen" href="#!">2 days ago</a>
</div>
<div class="organic-job even" data-url="spl/civil-engineer-aaaabbbbccccdddd1111222233334444?imo=2">
  <h2 class="title">
    <a class="lk-job-title" href="#!">Civil Engineer</a>
  </h2>
  <a class="employer lk-employer" href="#!">Build Co</a>
  <span class="description">Bridge construction and design.</span>
  <a class="lk lastseen" href="#!">1 day ago</a>
</div>
</body></html>
"""

SAMPLE_SPL_HTML = """
<html><body>
<main class="container-fluid">
<div class="col-xl-7 col-sm-7 col-12 description">
<div class="short-text">
<p>We are looking for a Backend Developer with 3-5 years of Python experience.</p>
<p>You will work with Django, PostgreSQL, and AWS.</p>
</div>
</div>
</main>
</body></html>
"""


def test_fetch_results_page_returns_jobs():
    from scraper import fetch_results_page
    config = {"scraper": {"delay_min": 0, "delay_max": 0, "respect_robots_txt": False}}
    pw_page = MagicMock()
    pw_page.goto.return_value = MagicMock(status=200)
    pw_page.content.return_value = SAMPLE_RESULTS_HTML
    pw_page.url = "https://www.eluta.ca/search?q=software+engineer"
    jobs = fetch_results_page(1, "software engineer", config, pw_page)
    assert len(jobs) == 2
    assert jobs[0]["title"] == "Backend Developer"
    assert jobs[0]["company"] == "Acme Corp"
    assert jobs[0]["job_id"] == "d46b145bbcc78f3ebfdc6d584e74f6e7"
    assert jobs[0]["slug"] == "spl/backend-developer-d46b145bbcc78f3ebfdc6d584e74f6e7?imo=1"
    assert jobs[0]["url"] == "https://www.eluta.ca/spl/backend-developer-d46b145bbcc78f3ebfdc6d584e74f6e7"


def test_fetch_results_page_empty_returns_empty_list():
    from scraper import fetch_results_page
    config = {"scraper": {"delay_min": 0, "delay_max": 0, "respect_robots_txt": False}}
    pw_page = MagicMock()
    pw_page.goto.return_value = MagicMock(status=200)
    pw_page.content.return_value = "<html><body></body></html>"
    pw_page.url = "https://www.eluta.ca/search?q=software+engineer"
    jobs = fetch_results_page(1, "software engineer", config, pw_page)
    assert jobs == []


def test_fetch_full_jd_returns_text():
    from scraper import fetch_full_jd
    config = {"scraper": {"delay_min": 0, "delay_max": 0}}
    pw_page = MagicMock()
    pw_page.goto.return_value = MagicMock(status=200)
    pw_page.content.return_value = SAMPLE_SPL_HTML
    text = fetch_full_jd("spl/backend-developer-abc123def456?imo=1", config, pw_page)
    assert "Python" in text
    assert "Django" in text
    assert "3-5 years" in text


def test_fetch_full_jd_returns_empty_on_parse_failure():
    from scraper import fetch_full_jd
    config = {"scraper": {"delay_min": 0, "delay_max": 0}}
    pw_page = MagicMock()
    pw_page.goto.return_value = MagicMock(status=200)
    pw_page.content.return_value = "<html><body><p>No description here.</p></body></html>"
    text = fetch_full_jd("spl/some-job?imo=1", config, pw_page)
    assert text == ""


def test_fetch_results_page_raises_on_http_error():
    from scraper import fetch_results_page
    from playwright.sync_api import Error as PlaywrightError
    config = {"scraper": {"delay_min": 0, "delay_max": 0, "respect_robots_txt": False}}
    pw_page = MagicMock()
    pw_page.goto.side_effect = PlaywrightError("net::ERR_CONNECTION_REFUSED")
    with pytest.raises(PlaywrightError):
        fetch_results_page(1, "software engineer", config, pw_page)


# ---------------------------------------------------------------------------
# TASK 10: Main Pipeline Tests
# ---------------------------------------------------------------------------

def _make_full_config():
    return {
        "sites": {"eluta": {"enabled": True, "query": "software engineer", "max_pages": 2}},
        "categories": {
            "backend": ["backend developer"],
            "general_swe": [],
        },
        "filters": {
            "seniority_blocklist": [" lead"],
            "non_technical_blocklist": ["millwright", "civil engineer"],
        },
        "classifier": {
            "confidence_threshold": 0.60,
            "claude_model": "claude-haiku-4-5-20251001",
            "max_few_shot_examples": 15,
        },
        "scraper": {"delay_min": 0, "delay_max": 0, "respect_robots_txt": False, "cutoff_days": 0},
    }


def test_pipeline_accepts_relevant_job():
    from scraper import run_scrape

    config = _make_full_config()
    feedback = {"decisions": [], "ambiguous_titles": []}

    page1_jobs = [
        {"title": "Backend Developer", "company": "Acme", "snippet": "Python APIs",
         "date_posted": "1 day ago", "job_id": "aaa111", "slug": "spl/backend-aaa111?imo=1",
         "url": "https://www.eluta.ca/spl/aaa111"},
    ]

    with patch("scraper.fetch_results_page") as mock_page, \
         patch("scraper.fetch_full_jd") as mock_jd, \
         patch("scraper._check_robots"), \
         patch("scraper.sync_playwright"):
        # Page 1 has 1 job, page 2 is empty → stops
        mock_page.side_effect = [page1_jobs, []]
        mock_jd.return_value = "3-5 years Python experience with Django and AWS."

        accepted, review, filtered, _, _, _ = run_scrape(config, feedback)

    assert len(accepted) == 1
    assert accepted[0]["title"] == "Backend Developer"
    assert accepted[0]["category"] == "backend"
    assert accepted[0]["yoe_required"] == "2-3"  # lower bound of "3-5 years" → _categorize_yoe(3) → "2-3"


def test_pipeline_filters_non_technical():
    from scraper import run_scrape

    config = _make_full_config()
    feedback = {"decisions": [], "ambiguous_titles": []}

    page1_jobs = [
        {"title": "Civil Engineer", "company": "Build Co", "snippet": "Bridges",
         "date_posted": "1 day ago", "job_id": "bbb222", "slug": "spl/civil-bbb222?imo=1",
         "url": "https://www.eluta.ca/spl/bbb222"},
    ]

    with patch("scraper.fetch_results_page") as mock_page, \
         patch("scraper.fetch_full_jd") as mock_jd, \
         patch("scraper._check_robots"), \
         patch("scraper.sync_playwright"):
        mock_page.side_effect = [page1_jobs, []]
        mock_jd.return_value = ""

        accepted, review, filtered, _, _, _ = run_scrape(config, feedback)

    assert len(accepted) == 0
    assert len(filtered) == 1
    assert filtered[0]["title"] == "Civil Engineer"


def test_pipeline_deduplicates_within_run():
    from scraper import run_scrape

    config = _make_full_config()
    feedback = {"decisions": [], "ambiguous_titles": []}

    same_job = {"title": "Backend Developer", "company": "Acme", "snippet": "Python",
                "date_posted": "1 day ago", "job_id": "aaa111", "slug": "spl/backend-aaa111?imo=1",
                "url": "https://www.eluta.ca/spl/aaa111"}

    with patch("scraper.fetch_results_page") as mock_page, \
         patch("scraper.fetch_full_jd") as mock_jd, \
         patch("scraper._check_robots"), \
         patch("scraper.sync_playwright"):
        # Same job appears on page 1 and page 2
        mock_page.side_effect = [[same_job], [same_job]]
        mock_jd.return_value = "2 years Python experience."

        accepted, review, filtered, _, _, _ = run_scrape(config, feedback)

    assert len(accepted) == 1  # deduplicated


def test_pipeline_stops_at_max_pages():
    from scraper import run_scrape

    config = _make_full_config()
    config["sites"]["eluta"]["max_pages"] = 2
    feedback = {"decisions": [], "ambiguous_titles": []}

    infinite_page = [
        {"title": "Backend Developer", "company": "Acme", "snippet": "Python",
         "date_posted": "1 day ago", "job_id": f"id{i}", "slug": f"spl/backend-id{i}?imo=1",
         "url": f"https://www.eluta.ca/spl/id{i}"}
        for i in range(10)
    ]

    with patch("scraper.fetch_results_page") as mock_page, \
         patch("scraper.fetch_full_jd") as mock_jd, \
         patch("scraper._check_robots"), \
         patch("scraper.sync_playwright"):
        mock_page.return_value = infinite_page  # always returns results
        mock_jd.return_value = "2 years Python experience."

        accepted, review, filtered, _, pages, _ = run_scrape(config, feedback)

    assert pages == 2  # stopped at max_pages, not from empty page


def test_pipeline_filters_feedback_rejected_title():
    from scraper import run_scrape

    config = _make_full_config()
    feedback = {
        "decisions": [
            {"title": "Backend Developer", "relevant": False, "category": None,
             "reason": "Not actually a dev role", "source": "review"}
        ],
        "ambiguous_titles": []
    }

    page1 = [
        {"title": "Backend Developer", "company": "Acme", "snippet": "Python",
         "date_posted": "1 day ago", "job_id": "aaa111", "slug": "spl/backend-aaa111?imo=1",
         "url": "https://www.eluta.ca/spl/aaa111"},
    ]

    with patch("scraper.fetch_results_page") as mock_page, \
         patch("scraper.fetch_full_jd") as mock_jd, \
         patch("scraper._check_robots"), \
         patch("scraper.sync_playwright"):
        mock_page.side_effect = [page1, []]
        mock_jd.return_value = "Some description."

        accepted, review, filtered, _, _, _ = run_scrape(config, feedback)

    assert len(accepted) == 0
    assert any(j["title"] == "Backend Developer" for j in filtered)


def test_pipeline_routes_low_confidence_to_review():
    from scraper import run_scrape

    config = _make_full_config()
    feedback = {"decisions": [], "ambiguous_titles": []}

    # Use a title that won't keyword-match → falls through to Claude
    page1_jobs = [
        {"title": "Technical Specialist", "company": "Corp", "snippet": "Some role",
         "date_posted": "1 day ago", "job_id": "ccc333", "slug": "spl/tech-ccc333?imo=1",
         "url": "https://www.eluta.ca/spl/ccc333"},
    ]

    mock_claude_result = {
        "relevant": True, "category": "general_swe", "confidence": 0.45,
        "yoe": "unknown", "flagged_for_review": True,
    }

    with patch("scraper.fetch_results_page") as mock_page, \
         patch("scraper.fetch_full_jd") as mock_jd, \
         patch("scraper.claude_classify") as mock_claude, \
         patch("scraper._check_robots"), \
         patch("scraper.sync_playwright"):
        mock_page.side_effect = [page1_jobs, []]
        mock_jd.return_value = "Some technical description."
        mock_claude.return_value = mock_claude_result

        accepted, review, filtered, _, _, _ = run_scrape(config, feedback)

    assert len(accepted) == 0
    assert len(review) == 1
    assert review[0]["title"] == "Technical Specialist"


def test_pipeline_pages_scraped_correct_on_early_empty_page():
    from scraper import run_scrape

    config = _make_full_config()
    config["sites"]["eluta"]["max_pages"] = 5
    feedback = {"decisions": [], "ambiguous_titles": []}

    page1_jobs = [
        {"title": "Backend Developer", "company": "Acme", "snippet": "Python",
         "date_posted": "1 day ago", "job_id": "aaa111", "slug": "spl/backend-aaa111?imo=1",
         "url": "https://www.eluta.ca/spl/aaa111"},
    ]

    with patch("scraper.fetch_results_page") as mock_page, \
         patch("scraper.fetch_full_jd") as mock_jd, \
         patch("scraper._check_robots"), \
         patch("scraper.sync_playwright"):
        # Page 1 has jobs, page 2 is empty → stops after 1 page
        mock_page.side_effect = [page1_jobs, []]
        mock_jd.return_value = "2 years Python experience."

        _, _, _, _, pages, _ = run_scrape(config, feedback)

    assert pages == 1  # only 1 page actually had results


def test_parse_days_ago():
    from scraper import _parse_days_ago
    assert _parse_days_ago("today") == 0
    assert _parse_days_ago("3 hours ago") == 0
    assert _parse_days_ago("yesterday") == 1
    assert _parse_days_ago("2 days ago") == 2
    assert _parse_days_ago("1 week ago") == 7
    assert _parse_days_ago("2 weeks ago") == 14
    assert _parse_days_ago("") == 0


def test_pipeline_stops_at_cutoff_days():
    from scraper import run_scrape

    config = _make_full_config()
    config["scraper"]["cutoff_days"] = 2
    feedback = {"decisions": [], "ambiguous_titles": []}

    page1 = [
        {"title": "Backend Developer", "company": "Acme", "snippet": "Python",
         "date_posted": "1 day ago", "job_id": "aaa111", "slug": "spl/backend-aaa111?imo=1",
         "url": "https://www.eluta.ca/spl/aaa111"},
    ]
    page2 = [
        {"title": "Backend Developer", "company": "Corp", "snippet": "Python",
         "date_posted": "5 days ago", "job_id": "bbb222", "slug": "spl/backend-bbb222?imo=1",
         "url": "https://www.eluta.ca/spl/bbb222"},
    ]

    with patch("scraper.fetch_results_page") as mock_page, \
         patch("scraper.fetch_full_jd") as mock_jd, \
         patch("scraper._check_robots"), \
         patch("scraper.load_seen_ids", return_value=set()), \
         patch("scraper.save_seen_ids"), \
         patch("scraper.sync_playwright"):
        mock_page.side_effect = [page1, page2]
        mock_jd.return_value = "2 years Python."

        accepted, _, _, _, pages, _ = run_scrape(config, feedback)

    assert pages == 1  # stopped before processing page 2's old jobs
    assert len(accepted) == 1  # only the 1-day-old job


def test_pipeline_skips_seen_ids_across_runs():
    from scraper import run_scrape

    config = _make_full_config()
    feedback = {"decisions": [], "ambiguous_titles": []}

    page1 = [
        {"title": "Backend Developer", "company": "Acme", "snippet": "Python",
         "date_posted": "1 day ago", "job_id": "aaa111", "slug": "spl/backend-aaa111?imo=1",
         "url": "https://www.eluta.ca/spl/aaa111"},
    ]

    with patch("scraper.fetch_results_page") as mock_page, \
         patch("scraper.fetch_full_jd") as mock_jd, \
         patch("scraper._check_robots"), \
         patch("scraper.sync_playwright"):
        mock_page.side_effect = [page1, []]
        mock_jd.return_value = "2 years Python."

        accepted, _, _, duplicate_count, _, _ = run_scrape(config, feedback, seen_ids={"aaa111"})

    assert len(accepted) == 0
    assert duplicate_count == 1  # skipped as already seen


# ---------------------------------------------------------------------------
# TASK 11: XLSX Exporter Tests
# ---------------------------------------------------------------------------

def _make_accepted_jobs():
    return [
        {
            "job_id": "abc123", "title": "Backend Developer", "company": "Acme Corp",
            "date_posted": "1 day ago", "category": "backend", "yoe_required": "2-3",
            "url": "https://www.eluta.ca/spl/abc123", "confidence": None,
        },
        {
            "job_id": "def456", "title": "ML Engineer", "company": "AI Co",
            "date_posted": "3 days ago", "category": "ai_ml", "yoe_required": "4-5",
            "url": "https://www.eluta.ca/spl/def456", "confidence": 0.87,
        },
    ]


def test_write_accepted_xlsx_creates_file(tmp_path):
    from scraper import write_accepted_xlsx
    out = tmp_path / "jobs.xlsx"
    write_accepted_xlsx(_make_accepted_jobs(), str(out))
    assert out.exists()


def test_write_accepted_xlsx_has_correct_headers(tmp_path):
    from scraper import write_accepted_xlsx
    out = tmp_path / "jobs.xlsx"
    write_accepted_xlsx(_make_accepted_jobs(), str(out))
    wb = load_workbook(str(out))
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, 9)]
    assert "title" in headers
    assert "category" in headers
    assert "yoe_required" in headers
    assert "url" in headers


def test_write_accepted_xlsx_has_correct_row_count(tmp_path):
    from scraper import write_accepted_xlsx
    out = tmp_path / "jobs.xlsx"
    write_accepted_xlsx(_make_accepted_jobs(), str(out))
    wb = load_workbook(str(out))
    ws = wb.active
    # 1 header row + 2 data rows
    assert ws.max_row == 3


def test_write_review_xlsx_has_confirm_column(tmp_path):
    from scraper import write_review_xlsx
    out = tmp_path / "review.xlsx"
    write_review_xlsx(_make_accepted_jobs(), str(out))
    wb = load_workbook(str(out))
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, 12)]
    assert "confirm" in headers
    assert "reason" in headers


def test_write_accepted_xlsx_url_is_hyperlink(tmp_path):
    from scraper import write_accepted_xlsx, _ACCEPTED_COLUMNS
    out = tmp_path / "jobs.xlsx"
    write_accepted_xlsx(_make_accepted_jobs(), str(out))
    wb = load_workbook(str(out))
    ws = wb.active
    url_col = _ACCEPTED_COLUMNS.index("url") + 1  # 1-based
    assert ws.cell(2, url_col).hyperlink is not None


def test_write_accepted_xlsx_rows_are_color_coded(tmp_path):
    from scraper import write_accepted_xlsx, _ACCEPTED_COLUMNS, CATEGORY_COLORS
    out = tmp_path / "jobs.xlsx"
    write_accepted_xlsx(_make_accepted_jobs(), str(out))
    wb = load_workbook(str(out))
    ws = wb.active
    title_col = _ACCEPTED_COLUMNS.index("title") + 1
    row2_color = ws.cell(2, title_col).fill.fgColor.rgb  # backend row
    row3_color = ws.cell(3, title_col).fill.fgColor.rgb  # ai_ml row
    assert row2_color.endswith(CATEGORY_COLORS["backend"])
    assert row3_color.endswith(CATEGORY_COLORS["ai_ml"])


# ---------------------------------------------------------------------------
# TASK 12: Filtered JSON Exporter Tests
# ---------------------------------------------------------------------------

def test_write_filtered_json_creates_file(tmp_path):
    from scraper import write_filtered_json
    jobs = [
        {"job_id": "abc123", "title": "Civil Engineer", "company": "Build Co",
         "date_posted": "1 day ago", "snippet": "Bridges and roads.",
         "filter_reason": "non-technical blocklist match: 'civil engineer'",
         "url": "https://www.eluta.ca/spl/abc123"},
    ]
    out = tmp_path / "filtered.json"
    write_filtered_json(jobs, str(out))
    assert out.exists()


def test_write_filtered_json_correct_structure(tmp_path):
    from scraper import write_filtered_json
    jobs = [
        {"job_id": "abc123", "title": "Civil Engineer", "company": "Build Co",
         "date_posted": "1 day ago", "snippet": "Bridges.",
         "filter_reason": "non-technical blocklist match: 'civil engineer'",
         "url": "https://www.eluta.ca/spl/abc123"},
    ]
    out = tmp_path / "filtered.json"
    write_filtered_json(jobs, str(out))
    data = json.loads(out.read_text())
    assert isinstance(data, list)
    assert data[0]["title"] == "Civil Engineer"
    assert "filter_reason" in data[0]


# ---------------------------------------------------------------------------
# TASK 13: Feedback Ingester Tests
# ---------------------------------------------------------------------------

def test_ingest_feedback_from_review_xlsx(tmp_path):
    from scraper import write_review_xlsx, ingest_feedback
    # Create a review file with user decisions filled in
    jobs = [
        {"job_id": "abc123", "title": "Platform Engineer", "company": "Acme",
         "date_posted": "1 day ago", "category": "cloud_devops", "yoe_required": "2-3",
         "url": "https://www.eluta.ca/spl/abc123", "confidence": 0.55,
         "confirm": "yes", "reason": "Uses AWS and Kubernetes"},
        {"job_id": "def456", "title": "Project Coordinator", "company": "Corp",
         "date_posted": "2 days ago", "category": "general_swe", "yoe_required": "unknown",
         "url": "https://www.eluta.ca/spl/def456", "confidence": 0.50,
         "confirm": "no", "reason": "Not a technical role"},
    ]
    review_path = tmp_path / "review.xlsx"
    write_review_xlsx(jobs, str(review_path))

    feedback = {"decisions": [], "ambiguous_titles": []}
    updated = ingest_feedback(str(review_path), feedback)

    assert len(updated["decisions"]) == 2
    confirmed = next(d for d in updated["decisions"] if d["title"] == "Platform Engineer")
    assert confirmed["relevant"] is True
    assert confirmed["category"] == "cloud_devops"

    rejected = next(d for d in updated["decisions"] if d["title"] == "Project Coordinator")
    assert rejected["relevant"] is False


def test_ingest_feedback_from_dispute_json(tmp_path):
    from scraper import ingest_feedback
    disputed = [
        {
            "job_id": "abc123",
            "title": "Controls Engineer",
            "company": "Acme",
            "date_posted": "1 day ago",
            "snippet": "Python and PLC experience required.",
            "filter_reason": "non-technical blocklist match: 'engineer'",
            "url": "https://www.eluta.ca/spl/abc123",
            "dispute": True,
            "reason": "Software controls role — Python and PLC",
        }
    ]
    json_path = tmp_path / "filtered.json"
    json_path.write_text(json.dumps(disputed))

    feedback = {"decisions": [], "ambiguous_titles": []}
    updated = ingest_feedback(str(json_path), feedback)

    # Title added to ambiguous list (goes to Claude next run)
    assert "controls engineer" in [t.lower() for t in updated["ambiguous_titles"]]
    # Also added as positive few-shot example
    assert any(d["title"] == "Controls Engineer" and d["relevant"] is True
               for d in updated["decisions"])


def test_ingest_feedback_skips_json_without_dispute_flag(tmp_path):
    from scraper import ingest_feedback
    jobs = [
        {"job_id": "abc123", "title": "Civil Engineer", "filter_reason": "non-technical",
         "snippet": "", "company": "", "date_posted": "", "url": ""}
        # No "dispute" key
    ]
    json_path = tmp_path / "filtered.json"
    json_path.write_text(json.dumps(jobs))

    feedback = {"decisions": [], "ambiguous_titles": []}
    updated = ingest_feedback(str(json_path), feedback)
    assert len(updated["decisions"]) == 0
    assert len(updated["ambiguous_titles"]) == 0


# ---------------------------------------------------------------------------
# TASK 15: End-to-End Smoke Test
# ---------------------------------------------------------------------------

def test_end_to_end_smoke(tmp_path):
    """
    Full pipeline smoke test with mocked HTTP and Claude.
    Verifies XLSX + JSON files are created and contain expected data.
    """
    from scraper import run_scrape, write_accepted_xlsx, write_review_xlsx, write_filtered_json

    config = _make_full_config()
    feedback = {"decisions": [], "ambiguous_titles": []}

    page1 = [
        {"title": "Backend Developer", "company": "Acme", "snippet": "Python APIs",
         "date_posted": "1 day ago", "job_id": "aaa111", "slug": "spl/backend-aaa111?imo=1",
         "url": "https://www.eluta.ca/spl/aaa111"},
        {"title": "Civil Engineer", "company": "Build Co", "snippet": "Bridges",
         "date_posted": "2 days ago", "job_id": "bbb222", "slug": "spl/civil-bbb222?imo=2",
         "url": "https://www.eluta.ca/spl/bbb222"},
        {"title": "Technical Specialist", "company": "TechCo", "snippet": "Support role",
         "date_posted": "1 day ago", "job_id": "ccc333", "slug": "spl/tech-ccc333?imo=3",
         "url": "https://www.eluta.ca/spl/ccc333"},
    ]

    mock_claude_response = MagicMock()
    mock_claude_response.content = [MagicMock(
        text='{"relevant": true, "category": "general_swe", "confidence": 0.55, "yoe": "2-3"}'
    )]

    with patch("scraper.fetch_results_page") as mock_page, \
         patch("scraper.fetch_full_jd") as mock_jd, \
         patch("scraper._check_robots"), \
         patch("scraper.anthropic.Anthropic") as MockClient, \
         patch("scraper.sync_playwright"):

        mock_page.side_effect = [page1, []]
        mock_jd.return_value = "3-5 years of Python, Django, REST API experience."
        MockClient.return_value.messages.create.return_value = mock_claude_response

        accepted, review, filtered, _, _, _ = run_scrape(config, feedback)

    # Backend Developer: keyword match → accepted
    assert any(j["title"] == "Backend Developer" for j in accepted)
    # Civil Engineer: hard filter → filtered
    assert any(j["title"] == "Civil Engineer" for j in filtered)
    # Technical Specialist: Claude confidence 0.55 < 0.60 → review
    assert any(j["title"] == "Technical Specialist" for j in review)

    # Write outputs and verify files
    write_accepted_xlsx(accepted, str(tmp_path / "jobs.xlsx"))
    write_review_xlsx(review, str(tmp_path / "review.xlsx"))
    write_filtered_json(filtered, str(tmp_path / "filtered.json"))

    assert (tmp_path / "jobs.xlsx").exists()
    assert (tmp_path / "review.xlsx").exists()
    assert (tmp_path / "filtered.json").exists()
